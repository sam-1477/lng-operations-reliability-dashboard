"""
Build an Excel workbook skeleton for the LNG Operations Technical Reliability Dashboard.

The workbook is a starter dashboard for portfolio demonstration. It uses public
industrial datasets adapted into LNG-style reliability categories. It does not
use ExxonMobil or PNG LNG operating data.

Inputs:
- data/processed/azure_machine_reliability_summary.csv
- data/processed/phmsa_incident_risk_summary.csv
- samples/metropt3/metropt3_compressor_sample.csv

Outputs:
- excel/lng_operations_reliability_dashboard.xlsx
- reports/excel_dashboard_build_report.md

Run from the project root:
    python scripts\\build_excel_dashboard_skeleton.py
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.legend import Legend
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.text import RichText
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.text import (
        CharacterProperties,
        Paragraph,
        ParagraphProperties,
        RichTextProperties,
    )
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as error:
    raise SystemExit(
        "This script requires pandas and openpyxl. Install them with:\n"
        "pip install pandas openpyxl\n\n"
        "Or install all project dependencies with:\n"
        "pip install -r requirements.txt"
    ) from error


PROJECT_ROOT = Path(__file__).resolve().parents[1]

AZURE_INPUT_PATH = (
    PROJECT_ROOT / "data" / "processed" / "azure_machine_reliability_summary.csv"
)
PHMSA_INPUT_PATH = (
    PROJECT_ROOT / "data" / "processed" / "phmsa_incident_risk_summary.csv"
)
METROPT3_INPUT_PATH = (
    PROJECT_ROOT / "samples" / "metropt3" / "metropt3_compressor_sample.csv"
)

WORKBOOK_OUTPUT_PATH = (
    PROJECT_ROOT / "excel" / "lng_operations_reliability_dashboard.xlsx"
)
REPORT_OUTPUT_PATH = PROJECT_ROOT / "reports" / "excel_dashboard_build_report.md"

REQUIRED_INPUTS = [
    AZURE_INPUT_PATH,
    PHMSA_INPUT_PATH,
    METROPT3_INPUT_PATH,
]

# Professional energy-sector workbook theme.
#
# These are generic dashboard colors inspired by major industrial/energy
# dashboards. They do not use ExxonMobil logos, trademarks, or official brand
# assets.
DEEP_NAVY = "0B1F3A"
ENERGY_RED = "C8102E"
DARK_GRAY = "333333"
MEDIUM_GRAY = "6B7280"
STEEL_BLUE = "486581"
MUTED_BLUE_GRAY = "78909C"
MUTED_TEAL = "477A7B"
DARK_SLATE = "37474F"
LIGHT_GRAY = "F2F4F7"
SECTION_GRAY = "E8EDF3"
WHITE = "FFFFFF"
PALE_RED = "FCE4E4"
PALE_BLUE = "EAF1F8"
PALE_AMBER = "FFF2CC"
PALE_GREEN = "E2F0D9"
BORDER_GRAY = "D9DEE5"

NAVY_FILL = PatternFill("solid", fgColor=DEEP_NAVY)
RED_FILL = PatternFill("solid", fgColor=ENERGY_RED)
LIGHT_RED_FILL = PatternFill("solid", fgColor=PALE_RED)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
SECTION_FILL = PatternFill("solid", fgColor=SECTION_GRAY)
BLUE_FILL = PatternFill("solid", fgColor=PALE_BLUE)
GREEN_FILL = PatternFill("solid", fgColor=PALE_GREEN)
AMBER_FILL = PatternFill("solid", fgColor=PALE_AMBER)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)

EQUIPMENT_CATEGORY_COLORS = [
    DEEP_NAVY,
    ENERGY_RED,
    STEEL_BLUE,
    MEDIUM_GRAY,
    MUTED_TEAL,
]
PRIORITY_CHART_COLORS = {
    "Critical": ENERGY_RED,
    "Medium": MEDIUM_GRAY,
    "High": STEEL_BLUE,
    "Low": DARK_SLATE,
}

WHITE_FONT = Font(color=WHITE, bold=True)
BODY_FONT = Font(color=DARK_GRAY)
BOLD_FONT = Font(color=DARK_GRAY, bold=True)
TITLE_FONT = Font(size=20, bold=True, color=DEEP_NAVY)
SUBTITLE_FONT = Font(size=11, italic=True, color=MEDIUM_GRAY)
SECTION_FONT = Font(size=12, bold=True, color=WHITE)
KPI_VALUE_FONT = Font(size=20, bold=True, color=DEEP_NAVY)
CRITICAL_VALUE_FONT = Font(size=20, bold=True, color=ENERGY_RED)
WARNING_FONT = Font(italic=True, color=ENERGY_RED)

AI_RECOMMENDATIONS_TITLE_ROW_HEIGHT = 28
AI_RECOMMENDATIONS_SUBTITLE_ROW_HEIGHT = 32
AI_RECOMMENDATIONS_HEADER_ROW_HEIGHT = 24
AI_RECOMMENDATIONS_BODY_ROW_HEIGHT = 60
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)
RED_BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color=ENERGY_RED),
)


def check_required_inputs() -> None:
    """Stop early with a clear message if a dashboard input file is missing."""
    missing_files = [path for path in REQUIRED_INPUTS if not path.exists()]

    if not missing_files:
        return

    print("Missing required dashboard input file(s):")
    for path in missing_files:
        print(f"- {path.relative_to(PROJECT_ROOT)}")

    raise SystemExit(
        "Run the data inspection and processing scripts first, then rerun this script."
    )


def read_dashboard_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Read the processed and sample CSV files used by the workbook."""
    print("Reading dashboard input files...")
    azure = pd.read_csv(AZURE_INPUT_PATH)
    phmsa = pd.read_csv(PHMSA_INPUT_PATH)
    metropt3 = pd.read_csv(METROPT3_INPUT_PATH)
    return azure, phmsa, metropt3


def safe_sheet_title(title: str) -> str:
    """Keep worksheet titles within Excel's 31-character limit."""
    return title[:31]


def apply_title(
    ws, title: str, subtitle: str | None = None, merge_to_column: int = 4
) -> None:
    """Add a consistent title block to a worksheet."""
    if merge_to_column > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_to_column)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    if subtitle:
        if merge_to_column > 1:
            ws.merge_cells(
                start_row=2, start_column=1, end_row=2, end_column=merge_to_column
            )
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 32


def hide_gridlines(ws) -> None:
    """Hide default worksheet gridlines where explicit styling is provided."""
    ws.sheet_view.showGridLines = False


def apply_header_style(ws, row_number: int, start_col: int, end_col: int) -> None:
    """Apply the standard navy header style to a row."""
    for col_number in range(start_col, end_col + 1):
        cell = ws.cell(row_number, col_number)
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def fill_area(ws, start_row: int, end_row: int, start_col: int, end_col: int, fill) -> None:
    """Fill a bounded visual area without touching unused worksheet space."""
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            cell.fill = fill


def apply_red_accent_line(ws, row_number: int, start_col: int, end_col: int) -> None:
    """Add a visible red accent line used below major worksheet headings."""
    for col_number in range(start_col, end_col + 1):
        ws.cell(row_number, col_number).fill = RED_FILL
    ws.row_dimensions[row_number].height = 5


def style_merged_range(
    ws,
    cell_range: str,
    value: object,
    fill,
    font: Font,
    alignment: Alignment | None = None,
) -> None:
    """Create and style a merged range for visible dashboard sections.

    OpenPyXL chart title formatting can be inconsistent across Excel versions,
    so worksheet-level merged headings give the dashboard a reliable visual
    structure even if Excel renders native chart titles differently.
    """
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = value
    top_left.fill = fill
    top_left.font = font
    top_left.alignment = alignment or Alignment(vertical="center", wrap_text=True)
    top_left.border = THIN_BORDER


def add_dashboard_section_heading(ws, cell_range: str, title: str) -> None:
    """Add a large navy worksheet heading above a dashboard chart."""
    style_merged_range(
        ws,
        cell_range,
        title,
        NAVY_FILL,
        Font(size=14, bold=True, color=WHITE),
        Alignment(horizontal="left", vertical="center", wrap_text=True),
    )


def set_column_widths(ws, min_width: int = 12, max_width: int = 42) -> None:
    """Auto-size columns with a cap so long notes do not make sheets unusable."""
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def style_range_as_grid(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """Apply simple borders and alignment to a cell range."""
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            cell.border = THIN_BORDER
            if not cell.font.bold:
                cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def make_chart_text(
    size_points: int,
    bold: bool = False,
    color: str = DARK_GRAY,
    rotation_degrees: int | None = None,
) -> RichText:
    """Return reusable chart text properties.

    OpenPyXL stores chart font sizes in hundredths of a point, so 12 pt is
    written as 1200. Keeping this in one helper makes chart styling easier for
    beginners to follow.
    """
    body_properties = None
    if rotation_degrees is not None:
        body_properties = RichTextProperties(rot=rotation_degrees * 60000)

    return RichText(
        bodyPr=body_properties,
        p=[
            Paragraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(
                        sz=size_points * 100,
                        b=bold,
                        solidFill=color,
                    )
                )
            )
        ]
    )


def style_chart(chart, title: str, width: float, height: float, accent_color: str) -> None:
    """Apply readable title, axis, legend, and series styling to a chart."""
    chart.title = title
    chart.width = width
    chart.height = height
    chart.style = 10

    # Make chart titles larger and bold. The try/except keeps the workbook build
    # robust if a future openpyxl version changes the chart text internals.
    try:
        chart.title.tx.rich.p[0].pPr = ParagraphProperties(
            defRPr=CharacterProperties(
                sz=1400,
                b=True,
                solidFill=DEEP_NAVY,
            )
        )
        chart.title.tx.rich.p[0].r[0].rPr = CharacterProperties(
            sz=1400,
            b=True,
            solidFill=DEEP_NAVY,
        )
    except (AttributeError, IndexError):
        pass

    chart.x_axis.txPr = make_chart_text(11, color=DARK_GRAY)
    chart.y_axis.txPr = make_chart_text(11, color=DARK_GRAY)
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None

    for axis in [chart.x_axis, chart.y_axis]:
        if axis.title is None:
            continue
        try:
            axis.title.tx.rich.p[0].pPr = ParagraphProperties(
                defRPr=CharacterProperties(
                    sz=1200,
                    b=True,
                    solidFill=DARK_GRAY,
                )
            )
            axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(
                sz=1200,
                b=True,
                solidFill=DARK_GRAY,
            )
        except (AttributeError, IndexError):
            pass

    if chart.legend is not None:
        chart.legend.position = "b"
        try:
            chart.legend.txPr = make_chart_text(10, color=DARK_GRAY)
        except AttributeError:
            pass

    for series in chart.series:
        series.graphicalProperties.solidFill = accent_color
        series.graphicalProperties.line.solidFill = accent_color


def enable_horizontal_chart_gridlines(chart) -> None:
    """Show horizontal major gridlines on a chart's value axis.

    The dashboard worksheet gridlines stay hidden. This helper only controls
    chart-internal major gridlines, which makes the PHMSA trend easier to read
    without adding visual clutter to the rest of the dashboard.
    """
    chart.y_axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(
            ln=LineProperties(solidFill=BORDER_GRAY)
        )
    )


def format_phmsa_year_axis(chart, year_count: int) -> None:
    """Force Microsoft 365 Excel to display PHMSA years as x-axis labels."""
    rotate_year_labels = year_count > 18

    # Axis titles are assigned before style_chart() so their title styling is
    # preserved. This helper only controls label placement and visibility.
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    # Keep the year labels visible below the columns and prevent Excel from
    # auto-skipping annual categories such as 2010, 2011, 2012, etc.
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.noMultiLvlLbl = True
    chart.x_axis.txPr = make_chart_text(
        10,
        color=DARK_GRAY,
        rotation_degrees=-45 if rotate_year_labels else 0,
    )

    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.txPr = make_chart_text(11, color=DARK_GRAY)


def format_number_columns(ws, header_row: int, data_start_row: int) -> None:
    """Apply basic number formats based on column names."""
    for cell in ws[header_row]:
        header = str(cell.value).lower() if cell.value is not None else ""
        column_letter = get_column_letter(cell.column)

        if any(word in header for word in ["count", "records", "year"]):
            number_format = "#,##0"
        elif "percent" in header or "score" in header:
            number_format = "0.0"
        elif any(word in header for word in ["cost", "downtime", "hours"]):
            number_format = "#,##0.0"
        elif any(word in header for word in ["avg", "temperature", "pressure", "current"]):
            number_format = "0.00"
        else:
            continue

        for row_number in range(data_start_row, ws.max_row + 1):
            ws[f"{column_letter}{row_number}"].number_format = number_format


def write_dataframe_sheet(ws, dataframe: pd.DataFrame, table_name: str) -> None:
    """Write a dataframe to a worksheet and format it as an Excel table."""
    hide_gridlines(ws)
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    style_range_as_grid(ws, 1, ws.max_row, 1, ws.max_column)
    apply_header_style(ws, row_number=1, start_col=1, end_col=ws.max_column)
    format_number_columns(ws, header_row=1, data_start_row=2)

    if ws.max_row >= 2 and ws.max_column >= 1:
        last_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=table_name, ref=f"A1:{last_cell}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    set_column_widths(ws)


def create_readme_sheet(wb: Workbook) -> None:
    """Create the README sheet with purpose, sources, and honesty wording."""
    ws = wb.create_sheet("README")
    hide_gridlines(ws)
    apply_title(ws, "LNG Operations Technical Reliability Dashboard", merge_to_column=2)

    rows = [
        ("Workbook purpose", "Starter Excel dashboard skeleton for operations technical reliability analysis."),
        (
            "Portfolio demonstration",
            "This workbook uses public industrial datasets adapted into LNG-style reliability categories.",
        ),
        (
            "Data integrity note",
            "No ExxonMobil or PNG LNG operating data is used. No confidential, proprietary, or licensee data is included.",
        ),
        ("Workbook status", "Dashboard starter/skeleton, not a final polished dashboard and not a live plant dashboard."),
        ("Data source 1", "Azure Predictive Maintenance"),
        ("Data source 2", "MetroPT-3 Compressor Dataset"),
        ("Data source 3", "PHMSA LNG/Gas Transmission Incident Data"),
        (
            "Generated on",
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ),
    ]

    ws.append([])
    ws.append(["Topic", "Details"])
    for topic, details in rows:
        ws.append([topic, details])

    style_range_as_grid(ws, 3, ws.max_row, 1, 2)
    apply_header_style(ws, row_number=3, start_col=1, end_col=2)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 95
    ws.freeze_panes = "A4"


def create_kpi_calculations_sheet(
    wb: Workbook, azure: pd.DataFrame, phmsa: pd.DataFrame, metropt3: pd.DataFrame
) -> dict[str, object]:
    """Create the KPI_Calculations sheet and return values for the dashboard."""
    ws = wb.create_sheet("KPI_Calculations")
    hide_gridlines(ws)
    apply_title(
        ws,
        "KPI Calculations",
        "Plain-language formula guidance plus current calculated values from dashboard-ready files.",
        merge_to_column=4,
    )
    apply_red_accent_line(ws, row_number=3, start_col=1, end_col=4)

    total_assets = len(azure)
    critical_assets = int((azure["maintenance_priority"] == "Critical").sum())
    high_priority_assets = int((azure["maintenance_priority"] == "High").sum())
    average_health = round(float(azure["asset_health_score"].mean()), 1)
    total_downtime = round(float(azure["estimated_downtime_hours"].sum()), 1)
    phmsa_records = int(phmsa["incident_count"].sum()) if "incident_count" in phmsa.columns else len(phmsa)
    metropt3_records = len(metropt3)

    kpis = [
        (
            "total_assets",
            "Count rows in Azure_Reliability_Summary",
            total_assets,
            "Count of Azure machines mapped into LNG-style assets.",
        ),
        (
            "critical_assets",
            "COUNTIF maintenance_priority = Critical",
            critical_assets,
            "Assets with Critical maintenance priority.",
        ),
        (
            "high_priority_assets",
            "COUNTIF maintenance_priority = High",
            high_priority_assets,
            "Assets with High maintenance priority.",
        ),
        (
            "average_asset_health_score",
            "AVERAGE asset_health_score",
            average_health,
            "Average 0-100 screening score from the Azure reliability summary.",
        ),
        (
            "total_estimated_downtime_hours",
            "SUM estimated_downtime_hours",
            total_downtime,
            "Failure count multiplied by the 8-hour downtime assumption.",
        ),
        (
            "phmsa_incident_risk_records",
            "SUM incident_count in PHMSA_Incident_Risk",
            phmsa_records,
            "Total PHMSA incident records represented by the grouped risk summary.",
        ),
        (
            "metropt3_sample_records",
            "Count rows in MetroPT3_Compressor_Sample",
            metropt3_records,
            "Rows in the GitHub-safe MetroPT-3 compressor sample.",
        ),
    ]

    kpi_header_row = ws.max_row + 1
    ws.append(["KPI", "Suggested Excel Formula", "Current Value", "Notes"])
    for row in kpis:
        ws.append(list(row))

    style_range_as_grid(ws, kpi_header_row, kpi_header_row + len(kpis), 1, 4)
    apply_header_style(ws, row_number=kpi_header_row, start_col=1, end_col=4)
    for row_number in range(kpi_header_row + 1, kpi_header_row + len(kpis) + 1):
        for col_number in range(1, 5):
            cell = ws.cell(row_number, col_number)
            cell.fill = LIGHT_FILL
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )
        ws.cell(row_number, 3).number_format = "#,##0.0"
        ws.cell(row_number, 3).font = BOLD_FONT
        ws.row_dimensions[row_number].height = 34

    # Helper table 1: maintenance priority count.
    priority_start_row = 16
    ws.cell(priority_start_row, 1, "Maintenance Priority")
    ws.cell(priority_start_row, 2, "Asset Count")
    priority_summary = (
        azure.groupby("maintenance_priority")
        .size()
        .reindex(["Critical", "High", "Medium", "Low"])
        .dropna()
        .astype(int)
    )
    for offset, (priority, count) in enumerate(priority_summary.items(), start=1):
        ws.cell(priority_start_row + offset, 1, priority)
        ws.cell(priority_start_row + offset, 2, int(count))

    # Helper table 2: average health by LNG equipment category.
    category_start_row = 16
    category_start_col = 4
    ws.cell(category_start_row, category_start_col, "LNG Equipment Category")
    ws.cell(category_start_row, category_start_col + 1, "Average Health Score")
    health_by_category = (
        azure.groupby("lng_equipment_category")["asset_health_score"]
        .mean()
        .round(1)
        .sort_values(ascending=True)
    )
    for offset, (category, score) in enumerate(health_by_category.items(), start=1):
        ws.cell(category_start_row + offset, category_start_col, category)
        ws.cell(category_start_row + offset, category_start_col + 1, float(score))

    # Helper table 3: PHMSA incident count by year.
    year_start_row = 16
    year_start_col = 8
    ws.cell(year_start_row, year_start_col, "Year")
    ws.cell(year_start_row, year_start_col + 1, "Incident Count")
    if "year" in phmsa.columns and "incident_count" in phmsa.columns:
        year_summary = (
            phmsa.assign(year=phmsa["year"].astype(str))
            .groupby("year")["incident_count"]
            .sum()
            .sort_index()
        )
    else:
        year_summary = pd.Series(dtype=float)

    for offset, (year, count) in enumerate(year_summary.items(), start=1):
        ws.cell(year_start_row + offset, year_start_col, year)
        ws.cell(year_start_row + offset, year_start_col + 1, int(count))

    style_range_as_grid(ws, priority_start_row, priority_start_row + len(priority_summary), 1, 2)
    apply_header_style(ws, row_number=priority_start_row, start_col=1, end_col=2)
    for row_number in range(priority_start_row + 1, priority_start_row + len(priority_summary) + 1):
        if ws.cell(row_number, 1).value == "Critical":
            ws.cell(row_number, 1).fill = LIGHT_RED_FILL
            ws.cell(row_number, 2).fill = LIGHT_RED_FILL
            ws.cell(row_number, 1).font = Font(color=ENERGY_RED, bold=True)
            ws.cell(row_number, 2).font = Font(color=ENERGY_RED, bold=True)

    style_range_as_grid(
        ws,
        category_start_row,
        category_start_row + len(health_by_category),
        category_start_col,
        category_start_col + 1,
    )
    apply_header_style(
        ws,
        row_number=category_start_row,
        start_col=category_start_col,
        end_col=category_start_col + 1,
    )
    style_range_as_grid(
        ws,
        year_start_row,
        year_start_row + len(year_summary),
        year_start_col,
        year_start_col + 1,
    )
    apply_header_style(
        ws,
        row_number=year_start_row,
        start_col=year_start_col,
        end_col=year_start_col + 1,
    )

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 72
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 18
    ws.row_dimensions[kpi_header_row].height = 28
    ws.freeze_panes = f"A{kpi_header_row + 1}"

    return {
        "total_assets": total_assets,
        "critical_assets": critical_assets,
        "high_priority_assets": high_priority_assets,
        "average_asset_health_score": average_health,
        "total_estimated_downtime_hours": total_downtime,
        "phmsa_incident_risk_records": phmsa_records,
        "metropt3_sample_records": metropt3_records,
        "priority_start_row": priority_start_row,
        "priority_row_count": len(priority_summary),
        "category_start_row": category_start_row,
        "category_row_count": len(health_by_category),
        "year_start_row": year_start_row,
        "year_row_count": len(year_summary),
    }


def create_dashboard_sheet(wb: Workbook, kpi_values: dict[str, object]) -> None:
    """Create a starter dashboard with KPI cards, chart placeholders, and charts."""
    ws = wb.create_sheet("Dashboard")
    hide_gridlines(ws)
    ws.sheet_view.zoomScale = 90
    fill_area(ws, 1, 74, 1, 15, WHITE_FILL)

    ws.merge_cells("A1:O1")
    ws["A1"] = "LNG Operations Technical Reliability Dashboard"
    ws["A1"].fill = NAVY_FILL
    ws["A1"].font = Font(size=24, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:O2")
    ws["A2"] = (
        "Portfolio demonstration using public industrial datasets adapted into "
        "LNG-style reliability categories"
    )
    ws["A2"].fill = SECTION_FILL
    ws["A2"].font = Font(size=11, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 26

    ws.merge_cells("A3:O3")
    ws["A3"] = (
        "Not ExxonMobil or PNG LNG operating data. Not a live plant dashboard. "
        "Not a formal mechanical integrity assessment."
    )
    ws["A3"].fill = LIGHT_RED_FILL
    ws["A3"].font = WARNING_FONT
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 26

    apply_red_accent_line(ws, row_number=4, start_col=1, end_col=15)

    kpi_cards = [
        ("Total assets", kpi_values["total_assets"], GREEN_FILL, KPI_VALUE_FONT, "#,##0"),
        (
            "Critical assets",
            kpi_values["critical_assets"],
            LIGHT_RED_FILL,
            CRITICAL_VALUE_FONT,
            "#,##0",
        ),
        (
            "High priority assets",
            kpi_values["high_priority_assets"],
            AMBER_FILL,
            KPI_VALUE_FONT,
            "#,##0",
        ),
        (
            "Average health score",
            kpi_values["average_asset_health_score"],
            BLUE_FILL,
            KPI_VALUE_FONT,
            "0.0",
        ),
        (
            "Total estimated downtime hours",
            kpi_values["total_estimated_downtime_hours"],
            BLUE_FILL,
            KPI_VALUE_FONT,
            "#,##0.0",
        ),
        (
            "Incident risk records",
            kpi_values["phmsa_incident_risk_records"],
            BLUE_FILL,
            KPI_VALUE_FONT,
            "#,##0",
        ),
        (
            "Compressor sample records",
            kpi_values["metropt3_sample_records"],
            BLUE_FILL,
            KPI_VALUE_FONT,
            "#,##0",
        ),
    ]

    card_positions = [
        (6, 1),
        (6, 5),
        (6, 9),
        (6, 13),
        (10, 1),
        (10, 5),
        (10, 9),
    ]
    for (label, value, fill, value_font, number_format), (start_row, start_col) in zip(
        kpi_cards, card_positions
    ):
        label_range = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(start_col + 2)}{start_row}"
        )
        value_range = (
            f"{get_column_letter(start_col)}{start_row + 1}:"
            f"{get_column_letter(start_col + 2)}{start_row + 1}"
        )
        note_range = (
            f"{get_column_letter(start_col)}{start_row + 2}:"
            f"{get_column_letter(start_col + 2)}{start_row + 2}"
        )
        style_merged_range(
            ws,
            label_range,
            label,
            NAVY_FILL,
            WHITE_FONT,
            Alignment(horizontal="left", vertical="center", wrap_text=True),
        )
        style_merged_range(
            ws,
            value_range,
            value,
            fill,
            value_font,
            Alignment(horizontal="left", vertical="center"),
        )
        ws[value_range.split(":")[0]].number_format = number_format
        style_merged_range(
            ws,
            note_range,
            "Current value",
            fill,
            Font(size=9, color=MEDIUM_GRAY),
            Alignment(horizontal="left", vertical="center"),
        )

    for row_number in [6, 10]:
        ws.row_dimensions[row_number].height = 26
        ws.row_dimensions[row_number + 1].height = 30
        ws.row_dimensions[row_number + 2].height = 20

    # Large worksheet headings are added above each chart because native chart
    # title font sizing is not always preserved by Excel/openpyxl round trips.
    add_dashboard_section_heading(ws, "A14:G14", "Maintenance Priority Count")
    add_dashboard_section_heading(ws, "H14:O14", "Asset Health By LNG Equipment Category")
    add_dashboard_section_heading(ws, "A38:O38", "PHMSA Incident Count By Year")
    ws.row_dimensions[14].height = 24
    ws.row_dimensions[38].height = 24

    # Keep the value-axis descriptor outside the chart so it cannot overlap
    # Microsoft 365 Excel's numeric y-axis labels.
    style_merged_range(
        ws,
        "A17:A31",
        "Asset Count",
        WHITE_FILL,
        Font(size=11, bold=True, color=DARK_GRAY),
        Alignment(horizontal="center", vertical="center", text_rotation=90),
    )

    kpi_ws = wb["KPI_Calculations"]

    # Chart 1: maintenance priority count.
    if kpi_values["priority_row_count"] > 0:
        # Use one series per priority so Excel can preserve distinct fills for
        # Critical, Medium, and any future priority categories reliably.
        priority_source_start_row = 16
        priority_source_start_col = 20  # T: helper source is outside the screenshot area.
        priority_count = int(kpi_values["priority_row_count"])
        priority_end_row = priority_source_start_row + priority_count
        priority_names = [
            kpi_ws.cell(int(kpi_values["priority_start_row"]) + offset, 1).value
            for offset in range(1, priority_count + 1)
        ]
        priority_values = [
            kpi_ws.cell(int(kpi_values["priority_start_row"]) + offset, 2).value
            for offset in range(1, priority_count + 1)
        ]
        ws.cell(priority_source_start_row, priority_source_start_col, "Priority")
        for series_offset, priority in enumerate(priority_names, start=1):
            ws.cell(
                priority_source_start_row,
                priority_source_start_col + series_offset,
                priority,
            )

        for offset, priority in enumerate(priority_names, start=1):
            dashboard_row = priority_source_start_row + offset
            ws.cell(dashboard_row, priority_source_start_col, priority)
            for series_offset, series_priority in enumerate(priority_names, start=1):
                value = (
                    priority_values[offset - 1]
                    if priority == series_priority
                    else None
                )
                ws.cell(
                    dashboard_row,
                    priority_source_start_col + series_offset,
                    value,
                )

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.gapWidth = 75
        chart.y_axis.title = None
        chart.x_axis.title = None
        categories = Reference(
            ws,
            min_col=priority_source_start_col,
            min_row=priority_source_start_row + 1,
            max_row=priority_end_row,
        )
        for series_offset in range(1, len(priority_names) + 1):
            data = Reference(
                ws,
                min_col=priority_source_start_col + series_offset,
                min_row=priority_source_start_row,
                max_row=priority_end_row,
            )
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        style_chart(
            chart,
            "Maintenance Priority Count",
            width=14,
            height=8.5,
            accent_color=ENERGY_RED,
        )
        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1
        chart.x_axis.noMultiLvlLbl = True
        chart.x_axis.txPr = make_chart_text(10, color=DARK_GRAY)
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.delete = False
        chart.y_axis.scaling.min = 0
        largest_priority_count = max(int(value or 0) for value in priority_values)
        chart.y_axis.scaling.max = max(
            20,
            ((largest_priority_count + 19) // 20 + 1) * 20,
        )
        chart.y_axis.majorUnit = 20
        chart.y_axis.txPr = make_chart_text(11, color=DARK_GRAY)
        for series, priority in zip(chart.series, priority_names):
            color = PRIORITY_CHART_COLORS.get(priority, STEEL_BLUE)
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color
        enable_horizontal_chart_gridlines(chart)
        ws.add_chart(chart, "B15")

    # Chart 2: asset health by equipment category.
    if kpi_values["category_row_count"] > 0:
        # Use one series per equipment category so Excel preserves a distinct
        # professional fill for every horizontal bar.
        category_source_start_row = 16
        category_source_start_col = 24  # X: helper source is outside the screenshot area.
        category_count = int(kpi_values["category_row_count"])
        category_end_row = category_source_start_row + category_count
        category_names = [
            kpi_ws.cell(int(kpi_values["category_start_row"]) + offset, 4).value
            for offset in range(1, category_count + 1)
        ]
        ws.cell(category_source_start_row, category_source_start_col, "Equipment Category")
        for series_offset, category in enumerate(category_names, start=1):
            ws.cell(
                category_source_start_row,
                category_source_start_col + series_offset,
                category,
            )

        for offset, category in enumerate(category_names, start=1):
            source_row = int(kpi_values["category_start_row"]) + offset
            dashboard_row = category_source_start_row + offset
            ws.cell(dashboard_row, category_source_start_col, category)
            for series_offset, series_category in enumerate(category_names, start=1):
                value = (
                    kpi_ws.cell(source_row, 5).value
                    if category == series_category
                    else None
                )
                ws.cell(
                    dashboard_row,
                    category_source_start_col + series_offset,
                    value,
                )

        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.gapWidth = 50
        chart.x_axis.title = None
        chart.y_axis.title = "Health Score"
        chart.legend = Legend(legendPos="r")
        categories = Reference(
            ws,
            min_col=category_source_start_col,
            min_row=category_source_start_row + 1,
            max_row=category_end_row,
        )
        for series_offset in range(1, len(category_names) + 1):
            data = Reference(
                ws,
                min_col=category_source_start_col + series_offset,
                min_row=category_source_start_row,
                max_row=category_end_row,
            )
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        style_chart(
            chart,
            "Asset Health By LNG Equipment Category",
            width=19,
            height=9.5,
            accent_color=DEEP_NAVY,
        )
        chart.x_axis.axPos = "l"
        chart.y_axis.axPos = "b"
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1
        chart.x_axis.noMultiLvlLbl = True
        chart.x_axis.txPr = make_chart_text(10, color=DARK_GRAY)
        chart.y_axis.txPr = make_chart_text(11, color=DARK_GRAY)
        chart.legend.position = "r"
        chart.legend.txPr = make_chart_text(9, color=DARK_GRAY)
        for series, color in zip(chart.series, EQUIPMENT_CATEGORY_COLORS):
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color
        ws.add_chart(chart, "H15")

    # Chart 3: PHMSA incident count by year.
    if kpi_values["year_row_count"] > 0:
        # Build a simple two-column chart source table on the Dashboard sheet.
        # Microsoft 365 can misinterpret a line chart sourced from another
        # worksheet, causing each year to appear as its own legend item. Keeping
        # this helper source as Year + Incident Count on the Dashboard sheet and
        # using one column-series makes the chart orientation explicit.
        phmsa_source_start_row = 38
        phmsa_source_start_col = 17
        ws.cell(phmsa_source_start_row, phmsa_source_start_col, "Year")
        ws.cell(phmsa_source_start_row, phmsa_source_start_col + 1, "Incident Count")
        for offset in range(1, int(kpi_values["year_row_count"]) + 1):
            ws.cell(
                phmsa_source_start_row + offset,
                phmsa_source_start_col,
                kpi_ws.cell(int(kpi_values["year_start_row"]) + offset, 8).value,
            )
            ws.cell(
                phmsa_source_start_row + offset,
                phmsa_source_start_col + 1,
                kpi_ws.cell(int(kpi_values["year_start_row"]) + offset, 9).value,
            )

        apply_header_style(
            ws,
            row_number=phmsa_source_start_row,
            start_col=phmsa_source_start_col,
            end_col=phmsa_source_start_col + 1,
        )
        style_range_as_grid(
            ws,
            phmsa_source_start_row,
            phmsa_source_start_row + int(kpi_values["year_row_count"]),
            phmsa_source_start_col,
            phmsa_source_start_col + 1,
        )
        ws.column_dimensions[get_column_letter(phmsa_source_start_col)].width = 12
        ws.column_dimensions[get_column_letter(phmsa_source_start_col + 1)].width = 16

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.gapWidth = 70
        chart.y_axis.title = "Incident Count"
        chart.x_axis.title = "Year"
        end = phmsa_source_start_row + int(kpi_values["year_row_count"])
        data = Reference(
            ws,
            min_col=phmsa_source_start_col + 1,
            min_row=phmsa_source_start_row,
            max_row=end,
        )
        categories = Reference(
            ws,
            min_col=phmsa_source_start_col,
            min_row=phmsa_source_start_row + 1,
            max_row=end,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        style_chart(
            chart,
            "PHMSA Incident Count By Year",
            width=28.5,
            height=13.5,
            accent_color=ENERGY_RED,
        )
        format_phmsa_year_axis(chart, int(kpi_values["year_row_count"]))
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.16,
                y=0.06,
                w=0.78,
                h=0.68,
            )
        )
        chart.legend = None
        enable_horizontal_chart_gridlines(chart)
        ws.add_chart(chart, "A42")

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for col in [1, 5, 9, 13]:
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 12
    # Dashboard screenshots should show the complete title and disclaimer area.
    ws.freeze_panes = None


def create_ai_recommendations_sheet(
    wb: Workbook, azure: pd.DataFrame, phmsa: pd.DataFrame
) -> None:
    """Create simple AI-style recommendations from available processed data."""
    ws = wb.create_sheet("AI_Recommendations")
    hide_gridlines(ws)
    apply_title(
        ws,
        "AI Recommendations",
        "Starter recommendation table for portfolio discussion. Review by a qualified engineer would be required before real use.",
        merge_to_column=4,
    )
    ws.row_dimensions[1].height = AI_RECOMMENDATIONS_TITLE_ROW_HEIGHT
    ws.row_dimensions[2].height = AI_RECOMMENDATIONS_SUBTITLE_ROW_HEIGHT
    apply_red_accent_line(ws, row_number=3, start_col=1, end_col=4)

    critical_count = int((azure["maintenance_priority"] == "Critical").sum())
    high_count = int((azure["maintenance_priority"] == "High").sum())
    incident_records = int(phmsa["incident_count"].sum()) if "incident_count" in phmsa.columns else len(phmsa)

    rows = [
        [
            "Asset reliability",
            f"{critical_count} assets are currently marked Critical.",
            "Review critical assets first and confirm the failure, error, and vibration drivers.",
            "Critical priority indicates repeated failures or poor public-data health score in the Azure reliability summary.",
        ],
        [
            "Maintenance planning",
            f"{high_count} assets are currently marked High priority.",
            "Plan inspection during the next maintenance window and review component replacement history.",
            "High priority assets may not require immediate shutdown, but they should not be ignored in a reliability review.",
        ],
        [
            "Incident risk awareness",
            f"PHMSA summary represents {incident_records} public incident records.",
            "Use PHMSA trends to discuss cause categories, consequence severity, and mechanical integrity context.",
            "PHMSA data supports risk awareness, not a formal integrity assessment for PNG LNG operations.",
        ],
        [
            "Compressor surveillance",
            "MetroPT-3 sample provides compressor pressure, oil temperature, current, and signal values.",
            "Monitor oil temperature, pressure trends, motor current, and operating-status signals.",
            "Public compressor trends demonstrate surveillance methods but are not LNG plant alarm limits.",
        ],
    ]

    header_row = ws.max_row + 1
    ws.append(["area", "finding", "recommendation", "engineering_rationale"])
    for row in rows:
        ws.append(row)

    style_range_as_grid(ws, header_row, ws.max_row, 1, 4)
    apply_header_style(ws, row_number=header_row, start_col=1, end_col=4)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 55
    ws.row_dimensions[header_row].height = AI_RECOMMENDATIONS_HEADER_ROW_HEIGHT
    for row_number in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[row_number].height = AI_RECOMMENDATIONS_BODY_ROW_HEIGHT
        for col_number in range(1, 5):
            cell = ws.cell(row_number, col_number)
            cell.fill = WHITE_FILL
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    ws.freeze_panes = f"A{header_row + 1}"


def create_assumptions_sheet(wb: Workbook) -> None:
    """Create the assumptions and limitations sheet."""
    ws = wb.create_sheet("Assumptions")
    hide_gridlines(ws)
    apply_title(
        ws,
        "Assumptions",
        "Public datasets are adapted into LNG-style reliability categories for portfolio learning.",
        merge_to_column=3,
    )

    rows = [
        (
            "Azure asset mapping",
            "Azure machine records are adapted into LNG-style asset categories.",
            "Use as a portfolio asset model, not a real facility register.",
        ),
        (
            "Operating-hour proxy",
            "Telemetry record count is used as an operating-hour proxy.",
            "MTBF-style values are approximate and should be labeled as estimates.",
        ),
        (
            "Downtime assumption",
            "Each failure is assumed to cause 8 downtime hours.",
            "Estimated downtime is a simple screening calculation.",
        ),
        (
            "Reliability KPIs",
            "MTBF-style and availability-style values are estimates for portfolio learning.",
            "Do not present as auditable plant reliability metrics.",
        ),
        (
            "MetroPT-3",
            "MetroPT-3 is a public compressor dataset adapted for LNG-style compressor surveillance.",
            "Trends are process-surveillance examples, not LNG alarm limits.",
        ),
        (
            "PHMSA",
            "PHMSA data is US public incident data used for risk-awareness context, not PNG LNG operating history.",
            "Use as incident awareness and mechanical integrity context only.",
        ),
        (
            "Assessment boundary",
            "This is not a formal mechanical integrity assessment or live plant dashboard.",
            "Workbook is a dashboard skeleton for GitHub, CV, and LinkedIn portfolio presentation.",
        ),
        (
            "Data confidentiality",
            "No ExxonMobil or PNG LNG operating data is used.",
            "All project claims must preserve this wording clearly.",
        ),
    ]

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Assumption", "Description", "Dashboard implication"])
    for row in rows:
        ws.append(row)

    style_range_as_grid(ws, header_row, ws.max_row, 1, 3)
    apply_header_style(ws, row_number=header_row, start_col=1, end_col=3)

    ws.append([])
    enhancement_section_row = ws.max_row + 1
    style_merged_range(
        ws,
        f"A{enhancement_section_row}:C{enhancement_section_row}",
        "Future Manual Enhancements",
        NAVY_FILL,
        SECTION_FONT,
        Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    ws.row_dimensions[enhancement_section_row].height = 24

    enhancement_header_row = enhancement_section_row + 1
    ws.append(["Enhancement", "Planned action", "Portfolio purpose"])
    enhancement_rows = [
        (
            "Slicers",
            "Add slicers for LNG equipment category, maintenance priority, source dataset, and year.",
            "Support interactive filtering during a future Excel presentation pass.",
        ),
        (
            "Formatting polish",
            "Review print area, spacing, number formats, and screenshot framing in Microsoft 365 Excel.",
            "Keep the recruiter-facing workbook presentation consistent and easy to scan.",
        ),
        (
            "Chart refinement",
            "Review chart labels, titles, axis scales, and category colors after opening the workbook in Excel.",
            "Confirm native Excel rendering matches the intended dashboard layout.",
        ),
        (
            "Validation",
            "Review assumptions, source descriptions, and portfolio wording before publication.",
            "Preserve clear boundaries around public datasets and estimated metrics.",
        ),
    ]
    for row in enhancement_rows:
        ws.append(row)

    style_range_as_grid(
        ws,
        enhancement_header_row,
        enhancement_header_row + len(enhancement_rows),
        1,
        3,
    )
    apply_header_style(ws, row_number=enhancement_header_row, start_col=1, end_col=3)
    for row_number in range(
        enhancement_header_row + 1,
        enhancement_header_row + len(enhancement_rows) + 1,
    ):
        for col_number in range(1, 4):
            ws.cell(row_number, col_number).fill = LIGHT_FILL
            ws.cell(row_number, col_number).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )
        ws.row_dimensions[row_number].height = 44

    set_column_widths(ws, max_width=65)
    ws.freeze_panes = f"A{header_row + 1}"


def create_workbook(azure: pd.DataFrame, phmsa: pd.DataFrame, metropt3: pd.DataFrame) -> None:
    """Create and save the Excel dashboard skeleton workbook."""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    create_readme_sheet(wb)

    azure_ws = wb.create_sheet("Azure_Reliability_Summary")
    write_dataframe_sheet(azure_ws, azure, "tblAzureReliabilitySummary")

    phmsa_ws = wb.create_sheet("PHMSA_Incident_Risk")
    write_dataframe_sheet(phmsa_ws, phmsa, "tblPHMSAIncidentRisk")

    metro_ws = wb.create_sheet("MetroPT3_Compressor_Sample")
    write_dataframe_sheet(metro_ws, metropt3, "tblMetroPT3CompressorSample")

    kpi_values = create_kpi_calculations_sheet(wb, azure, phmsa, metropt3)
    create_dashboard_sheet(wb, kpi_values)
    create_ai_recommendations_sheet(wb, azure, phmsa)
    create_assumptions_sheet(wb)

    # Keep the requested order even if sheets were created in helper functions.
    requested_order = [
        "README",
        "Azure_Reliability_Summary",
        "PHMSA_Incident_Risk",
        "MetroPT3_Compressor_Sample",
        "KPI_Calculations",
        "Dashboard",
        "AI_Recommendations",
        "Assumptions",
    ]
    wb._sheets = [wb[sheet_name] for sheet_name in requested_order]
    wb.active = requested_order.index("Dashboard")

    # Enforce reduced visual clutter after all sheets have been created.
    for worksheet in wb.worksheets:
        hide_gridlines(worksheet)

    WORKBOOK_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_OUTPUT_PATH)


def build_report(azure: pd.DataFrame, phmsa: pd.DataFrame, metropt3: pd.DataFrame) -> str:
    """Create the Excel dashboard build report."""
    return "\n".join(
        [
            "# Excel Dashboard Build Report",
            "",
            "## Workbook Created",
            "",
            f"- Workbook: `{WORKBOOK_OUTPUT_PATH.relative_to(PROJECT_ROOT).as_posix()}`",
            "- Status: polished starter/skeleton workbook, not a live plant dashboard",
            "- Purpose: operations technical reliability dashboard portfolio demonstration",
            "- Visual polish: visible worksheet-level chart headings, dashboard title band, KPI cards, gridline visibility, KPI table readability, and AI recommendation formatting were improved.",
            "- Theme: professional red/navy/white/light-gray energy-sector dashboard theme.",
            "- Branding boundary: no ExxonMobil logos, trademarks, or official brand assets were used.",
            "",
            "## Datasets Feeding The Workbook",
            "",
            f"- `data/processed/azure_machine_reliability_summary.csv` ({len(azure):,} rows)",
            f"- `data/processed/phmsa_incident_risk_summary.csv` ({len(phmsa):,} grouped rows)",
            f"- `samples/metropt3/metropt3_compressor_sample.csv` ({len(metropt3):,} rows)",
            "",
            "These are public industrial datasets adapted into LNG-style reliability "
            "categories. They are not ExxonMobil or PNG LNG operating data.",
            "",
            "## Sheets Created",
            "",
            "- `README`",
            "- `Azure_Reliability_Summary`",
            "- `PHMSA_Incident_Risk`",
            "- `MetroPT3_Compressor_Sample`",
            "- `KPI_Calculations`",
            "- `Dashboard`",
            "- `AI_Recommendations`",
            "- `Assumptions`",
            "",
            "## KPIs Included",
            "",
            "- Total assets",
            "- Critical assets",
            "- High priority assets",
            "- Average asset health score",
            "- Total estimated downtime hours",
            "- PHMSA incident risk records",
            "- MetroPT-3 sample records",
            "",
            "## Charts Included",
            "",
            "- Maintenance priority count, with Critical red and Medium muted gray as separate chart series",
            "- Asset health by LNG equipment category, shown as a horizontal multi-color bar chart with readable category labels and no obstructive category-axis title",
            "- PHMSA incident count by year, corrected as a single-series red column chart with Year as the x-axis category, visible x-axis year labels, and visible horizontal chart gridlines for readability",
            "",
            "Dashboard visual polish was applied with a large navy title band, a subtitle, a visible red disclaimer area, merged KPI cards, worksheet-level chart headings, clearer chart spacing, hidden worksheet gridlines, and red/navy accent colors. The Maintenance Priority Count chart uses separate series so Critical is red and Medium is muted gray, with readable priority labels and a compact legend. The Asset Health By LNG Equipment Category chart was changed to a horizontal bar chart so category labels remain readable, each category uses a distinct muted navy, red, steel, gray, or teal fill, and the obstructive Equipment Category axis title was removed. The PHMSA Incident Count By Year chart was corrected so Year is used as the x-axis category, Incident Count is one meaningful data series, and every annual x-axis label is requested for Microsoft 365 Excel display. The previous year-by-year legend issue was fixed by using a Dashboard-local two-column chart source and removing the legend. The PHMSA chart height and x-axis layout were adjusted so the year labels have more room below the red columns. Horizontal chart gridlines were enabled to make the year-to-year incident trend easier to read. Native chart title and axis text styling is also attempted through openpyxl, but the worksheet headings provide a reliable visible fallback in Excel.",
            "",
            "## Formatting Improvements",
            "",
            "- `KPI_Calculations` uses plain-text formula guidance instead of broken `#REF!` references.",
            "- `KPI_Calculations` body text contrast was improved with dark gray normal table text while retaining the navy header row, white bold header text, red accent line, light-gray calculation rows, borders, fixed widths, wrapped notes, and freeze panes.",
            "- `AI_Recommendations` body text contrast was improved with dark gray finding, recommendation, and engineering rationale text while retaining the navy header row, white bold header text, red accent line, requested column widths, and freeze panes below the header.",
            "- `AI_Recommendations` row spacing was reduced while keeping wrapped text readable: title row 28, subtitle row 32, header row 24, and body rows 60.",
            "- The workbook opens on the `Dashboard` sheet and hides worksheet gridlines to reduce visual clutter.",
            "- Maintenance Priority bars were given distinct colors through separate chart series: Critical uses energy red and Medium uses muted gray.",
            "- Maintenance Priority chart axis count visibility was improved with a 0-based y-axis, visible 20-unit major ticks, numeric labels, and horizontal gridlines.",
            "- Maintenance Priority Count y-axis label readability was fixed by removing the native chart title and placing a separate worksheet `Asset Count` label outside the chart; the y-axis numeric scale remains visible.",
            "- Asset Health equipment categories were given distinct colors through separate chart series using a consistent navy, red, steel blue, gray, and muted teal palette.",
            "- Asset Health chart legend/category readability was improved with a visible right-side legend and readable left-side category labels.",
            "- Obstructive x-axis label placement was fixed by removing the Asset Health `Equipment Category` axis title and leaving the category labels on the left category axis.",
            "- PHMSA chart layout was enlarged and adjusted so its Year and Incident Count axis labels remain readable.",
            "- Dashboard chart spacing was improved to prevent overlap with section headers and preserve clean screenshot framing.",
            "- Worksheet gridlines remain hidden; the PHMSA chart gridlines are chart-internal horizontal major gridlines only.",
            "- The PHMSA chart uses a clean column chart because Microsoft 365 was not reliably rendering the previous line-chart gridlines or legend orientation.",
            "- PHMSA chart x-axis year labels were made visible with bottom tick-label placement and no major tick-label skipping.",
            "- PHMSA chart height/layout was adjusted for readability; the x-axis title remains `Year` and the y-axis title remains `Incident Count`.",
            "- The former `Manual Excel Work Remaining` table was removed from the Dashboard and moved to the `Assumptions` sheet under `Future Manual Enhancements`, covering slicers, formatting polish, chart refinement, and validation.",
            "- Dashboard freeze panes were removed so the header no longer obstructs scrolling or screenshot capture.",
            "- Data sheets keep frozen headers and readable table formatting with consistent navy header styling where practical.",
            "",
            "## Assumptions",
            "",
            "- Azure machines are mapped into LNG-style asset categories.",
            "- Telemetry record count is used as an operating-hour proxy.",
            "- Each failure is assumed to cause 8 downtime hours.",
            "- MTBF-style and availability-style values are estimates for portfolio learning.",
            "- MetroPT-3 is adapted for LNG-style compressor surveillance.",
            "- PHMSA is public US incident data used for risk-awareness context.",
            "- The workbook is not a formal mechanical integrity assessment or live plant dashboard.",
            "",
            "## Future Manual Enhancements",
            "",
            "The former `Manual Excel Work Remaining` content was moved from the Dashboard to the `Assumptions` sheet under `Future Manual Enhancements`.",
            "- Slicers for maintenance priority, LNG equipment category, source dataset, and year.",
            "- Formatting polish for print area, spacing, number formats, and screenshot framing.",
            "- Chart refinement review for labels, titles, axis scales, and native Excel rendering.",
            "- Validation of assumptions and portfolio wording before publication.",
            "",
            "## Portfolio Story Support",
            "",
            "This workbook connects Azure reliability scoring, MetroPT-3 compressor "
            "surveillance, and PHMSA incident risk awareness into one operations "
            "technical reliability dashboard. It demonstrates data preparation, "
            "reliability thinking, risk-awareness communication, and Excel-based "
            "presentation using public industrial datasets adapted into LNG-style "
            "reliability categories. The workbook remains a portfolio demonstration "
            "using public datasets, not ExxonMobil or PNG LNG operating data.",
            "",
        ]
    )


def main() -> None:
    """Run the workbook build workflow."""
    print("Building Excel dashboard skeleton...")
    print(f"Project root: {PROJECT_ROOT}")

    check_required_inputs()
    azure, phmsa, metropt3 = read_dashboard_inputs()

    create_workbook(azure, phmsa, metropt3)
    REPORT_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_OUTPUT_PATH.write_text(build_report(azure, phmsa, metropt3), encoding="utf-8")

    print()
    print("Excel dashboard skeleton build complete.")
    print(f"Workbook written: {WORKBOOK_OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    print(f"Build report written: {REPORT_OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    print(f"Azure rows loaded: {len(azure):,}")
    print(f"PHMSA grouped rows loaded: {len(phmsa):,}")
    print(f"MetroPT-3 sample rows loaded: {len(metropt3):,}")


if __name__ == "__main__":
    main()
